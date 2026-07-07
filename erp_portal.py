"""
ERP Portal (Semester Readiness & Closure) — integrated into the OoA university app
as a Flask Blueprint mounted at /erp. Originally jain-portal-v2.
Uses its own MongoDB database (DB_NAME, default: semreadiness) on the same cluster.
"""
from flask import (Blueprint, render_template, request, jsonify, session, redirect,
                   url_for, send_file, flash, abort, send_from_directory, current_app)
from pymongo import MongoClient
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from bson import ObjectId
from dotenv import load_dotenv
import json, os, io, openpyxl, random, secrets
from datetime import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

load_dotenv()

erp_bp = Blueprint('erp', __name__)

ERP_UPLOAD_FOLDER = 'uploads'
ERP_MAX_CONTENT_LENGTH = 16 * 1024 * 1024  # informational; global cap is set by main app

# ══════════════════════════════════════════════════════════════════════
#  SINGLE SIGN-ON — the Jain Hub session powers the ERP portal.
#  Users logged into the university app are auto-authenticated here;
#  hub admins are automatically ERP admins. No separate login screens.
# ══════════════════════════════════════════════════════════════════════
@erp_bp.before_request
def _sso_bridge():
    hub_email = session.get('email')
    if hub_email:
        if session.get('user_email') != hub_email:
            session['user_email'] = hub_email
            session['user_name'] = (session.get('name')
                                    or hub_email.split('@')[0].replace('.', ' ').replace('_', ' ').title())
        try:
            if not users_col.find_one({'email': hub_email}, {'_id': 1}):
                users_col.insert_one({
                    'name': session['user_name'],
                    'email': hub_email,
                    'password': generate_password_hash(secrets.token_urlsafe(24)),  # unusable; SSO only
                    'department': session.get('user_department', ''),
                    'created_at': datetime.utcnow().isoformat(),
                    'first_time_login': True,
                    'sso': True
                })
        except Exception:
            pass
        if session.get('role') == 'admin' and not session.get('admin'):
            session['admin'] = True

@erp_bp.route('/manifest.json')
def serve_manifest():
    return send_from_directory('static/erp', 'manifest.json')

@erp_bp.route('/sw.js')
def serve_sw():
    return send_from_directory('static/erp', 'sw.js')

# MongoDB connection — sourced from .env (MONGO_URI + DB_NAME)
MONGO_URI = os.environ.get('MONGO_URI') or os.environ.get('MONGODB_URI')
DB_NAME = os.environ.get('DB_NAME', 'semreadiness')
client = MongoClient(MONGO_URI)
db = client[DB_NAME]
submissions_col = db['submissions']
faculty_submissions_col = db['faculty_submissions']  # Faculty's individual checklist submissions (for both readiness & closure)
users_col = db['users']
settings_col = db['settings']

ADMIN_USERNAME = os.environ.get('ADMIN_USERNAME', 'admin')
ADMIN_PASSWORD = os.environ.get('ADMIN_PASSWORD', 'admin2023')

DEPARTMENTS = [
    "Department of Computer Science and Engineering",
    "Department of Information Science and Engineering",
    "Department of Aerospace Engineering",
    "Department of Civil Engineering",
    "Department of Mechanical Engineering",
    "Department of Electrical and Electronics Engineering",
    "Department of Electronics and Communication Engineering",
    "Department of Food Technology",
    "Department of Humanities & Social Sciences",
    "Department of CERSSE",
    "Department of SSER",
    "Department of Jainology",
    "Department of Marine Science",
    "Department of Economics",
    "Department of Performing Arts and Cultural Studies",
    "Department of Languages",
    "Department of Journalism and Mass Communication",
    "Department of Law",
    "Department of Chemistry and Biochemistry",
    "Department of Biotechnology and Genetics",
    "Department of Microbiology and Botany",
    "Department of Data Analytics and Mathematical Science",
    "Department of Forensic Science",
    "Department of Physics and Electronics",
    "Department of Psychology and Allied Sciences",
    "Department of Allied Healthcare and Sciences",
    "Department of Computer Science and IT",
    "Department of Animation and Virtual Reality",
    "Department of Commerce",
    "Department of Management Studies",
    "Department of Design",
    "Department of Art and Design",
]

# ══════════════════════════════════════════════════════════════════
# READINESS — HOD CHECKLIST (HOD-only items, faculty items removed)
# ══════════════════════════════════════════════════════════════════
HOD_SECTIONS = [
    {"title": "Section 1: Curriculum & Course Matrix", "items": [
        {"id": "1", "text": "Previous semester's closing report insights incorporated"},
        {"id": "2", "text": "Course matrix verified (codes, credits, CBCS, LTPE) against BoS docs"},
        {"id": "3", "text": "Course matrix uploaded on ERP"},
        {"id": "4", "text": "Open Elective courses cross-checked with offering departments"},
        {"id": "5", "text": "Students given minimum 1 week to choose electives after orientation"},
        {"id": "6", "text": "Kochi / Online / ODL / other campus stakeholders included in planning"},
    ]},
    {"title": "Section 2: Faculty Allocation & Workload", "items": [
        {"id": "7", "text": "All courses have faculty assigned"},
        {"id": "8", "text": "Workload is fair and transparent across the department"},
        {"id": "9", "text": "Faculty informed of their course allocations formally"},
        {"id": "10", "text": "Faculty workload documented and assessed"},
    ]},
    {"title": "Section 3: TLEP Audit (HOD-level)", "items": [
        {"id": "12", "text": "HOD has reviewed / audited all TLEPs"},
        {"id": "13", "text": "Innovative pedagogies and assessments collated from TLEPs"},
        {"id": "14", "text": "Previous semester faculty feedback shared with faculty"},
    ]},
    {"title": "Section 4: Timetable & ERP", "items": [
        {"id": "15", "text": "Master timetable prepared"},
        {"id": "16", "text": "Master timetable uploaded on ERP"},
        {"id": "17", "text": "Course matrix created and approved in Faculty"},
    ]},
    {"title": "Section 5: LMS & Digital Readiness (HOD-level)", "items": [
        {"id": "19", "text": "All faculty trained on LMS platform"},
    ]},
    {"title": "Section 6: Students & Activities", "items": [
        {"id": "21", "text": "Co-curricular and extra-curricular activities planned and calendarised"},
        {"id": "22", "text": "Student Progression and Graduation report reviewed"},
        {"id": "23", "text": "Departmental Academic Achievement report prepared"},
    ]},
    {"title": "Section 7: Faculty Development & Research", "items": [
        {"id": "24", "text": "Faculty development / capacity building activities planned"},
        {"id": "25", "text": "Research activities (conferences, seminars) planned and calendarised"},
    ]},
    {"title": "Section 8: Infrastructure & Approvals", "items": [
        {"id": "26", "text": "Lab equipment, requirements and budget assessed"},
        {"id": "27", "text": "All necessary approvals obtained from University HO"},
        {"id": "28", "text": "Mid Sem Review Date"},
    ]},
]

# ══════════════════════════════════════════════════════════════════
# READINESS — FACULTY CHECKLIST (each faculty fills for their course)
# ══════════════════════════════════════════════════════════════════
FACULTY_SECTIONS = [
    {"title": "Section A: TLEP & Course Planning", "items": [
        {"id": "F1", "text": "TLEP submitted for the course"},
        {"id": "F2", "text": "Session plans prepared for entire semester"},
        {"id": "F3", "text": "Course objectives & outcomes mapped to PO/PSO"},
        {"id": "F4", "text": "Assessment plan finalized CIE/SEE components"},
    ]},
    {"title": "Section B: ERP & Documentation", "items": [
        {"id": "F5", "text": "Session plan uploaded on ERP"},
        {"id": "F6", "text": "Course material (notes/PPTs) prepared"},
        {"id": "F7", "text": "Reference textbooks & resources listed"},
    ]},
    {"title": "Section C: LMS & Digital Readiness", "items": [
        {"id": "F8", "text": "LMS course is live & accessible to students"},
        {"id": "F9", "text": "Week-1 content uploaded on LMS"},
        {"id": "F10", "text": "Discussion forums / assignments configured on LMS"},
    ]},
    {"title": "Section D: Pedagogy & Innovation", "items": [
        {"id": "F11", "text": "Innovative pedagogies planned (case studies, flipped classroom, etc.)"},
        {"id": "F12", "text": "Industry/practical examples integrated in course"},
    ]},
]

# ══════════════════════════════════════════════════════════════════
# CLOSURE — HOD REPORT SECTIONS (14 verticals, HOD-only narratives)
# ══════════════════════════════════════════════════════════════════
HOD_CLOSURE_SECTIONS = [
    {"id": "1",  "title": "Overview of the Semester",                                  "hint": "Brief narrative of how the semester progressed overall."},
    {"id": "2",  "title": "Objectives Planned — Were They Achieved?",                  "hint": "State the key objectives set at the beginning and whether each was met."},
    {"id": "3",  "title": "Adherence to Academic Calendar",                            "hint": "Highlight any deviations from the calendar and corrective actions taken."},
    {"id": "4",  "title": "Course Completion Summary",                                 "hint": "Summary of courses offered and completed. Flag any incomplete ones with reasons."},
    {"id": "5",  "title": "Student Performance Overview",                              "hint": "Average grades, top performers, notable academic achievements."},
    {"id": "6",  "title": "Faculty Contributions",                                     "hint": "Research published, conferences presented, awards received this semester."},
    {"id": "7",  "title": "Conferences, Workshops, Seminars & Guest Lectures",         "hint": "List all events conducted with dates and participants."},
    {"id": "8",  "title": "Department Events & Extra-curricular Activities",           "hint": "Overview of departmental events, fests, competitions etc."},
    {"id": "9",  "title": "Collaborations & Partnerships",                             "hint": "New MoUs or collaborations with Industry / Academia signed this semester."},
    {"id": "10", "title": "Academic Audits",                                           "hint": "Confirm that academic audits were conducted for all courses. Note any findings."},
    {"id": "11", "title": "CA Components — Entry & Approval",                          "hint": "Confirm CA marks are entered and approved in ERP for all courses."},
    {"id": "12", "title": "COLAB Files Review",                                        "hint": "Status of COLAB file review for this semester."},
    {"id": "13", "title": "COLAB — Previous Closing, Current Opening & Attainment",    "hint": "Confirm COLAB linkage: previous closing report reviewed, current opening report ready, CO attainment computed."},
    {"id": "14", "title": "HOD Additional Observations / Action Items for Next Semester","hint": "Any other noteworthy points or action items to carry forward."},
]

# ══════════════════════════════════════════════════════════════════
# CLOSURE — FACULTY CHECKLIST (each faculty fills for their course)
# Mirrors TEMPLATE_Closure_Checklist_-_Faculty_Odd_Semester.xlsx
# ══════════════════════════════════════════════════════════════════
FACULTY_CLOSURE_SECTIONS = [
    {"title": "Section A: ERP & Records", "items": [
        {"id": "FC1",  "text": "Attendance Entry in ERP with Analysis Report"},
        {"id": "FC2",  "text": "Continuous Assessment Marks Entry in ERP"},
        {"id": "FC3",  "text": "CA components mapped to COs / Bloom's Level"},
        {"id": "FC4",  "text": "Consolidated CA list available"},
        {"id": "FC5",  "text": "Assignment record maintained"},
    ]},
    {"title": "Section B: Syllabus & Teaching", "items": [
        {"id": "FC6",  "text": "Compliance to Teaching-Learning-Evaluation Plan (TLEP)"},
        {"id": "FC7",  "text": "Completion of Syllabus"},
        {"id": "FC8",  "text": "Innovative teaching methods adapted"},
    ]},
    {"title": "Section C: Learning Materials", "items": [
        {"id": "FC9",  "text": "Availability of Lecture Notes / PPTs"},
        {"id": "FC10", "text": "Sample Lecture Notes / PPT uploaded"},
        {"id": "FC11", "text": "Experiential Learning record with rubrics for Evaluation"},
        {"id": "FC12", "text": "Sample Experiential Learning activity with rubrics uploaded"},
    ]},
    {"title": "Section D: Evaluation", "items": [
        {"id": "FC13", "text": "Innovative Evaluation Strategy used"},
        {"id": "FC14", "text": "Availability of CO-PO (and PSO, if applicable) mapping"},
        {"id": "FC15", "text": "Course End Survey Conducted"},
    ]},
    {"title": "Section E: Student Support", "items": [
        {"id": "FC16", "text": "Slow learner / advance learner list prepared"},
        {"id": "FC17", "text": "Record of remedial classes maintained"},
        {"id": "FC18", "text": "Record of Extra lectures for Guided Self Study (GSS)"},
        {"id": "FC19", "text": "Lab Manuals available (wherever lab is a part of CA)"},
        {"id": "FC20", "text": "Result analysis with backlog list (course-wise)"},
    ]},
    {"title": "Section F: Mentoring & Projects", "items": [
        {"id": "FC21", "text": "Mentoring Report prepared"},
        {"id": "FC22", "text": "TD-PCL Report (batches, students, faculty, project title, progress)"},
        {"id": "FC23", "text": "Internship Report — Groups"},
    ]},
]

# ── Mail Helper ──────────────────────────────────────────

def get_base_url():
    # If in a request context, get the actual current domain
    try:
        from flask import has_request_context, request
        if has_request_context():
            return request.host_url
    except Exception:
        pass

    # Fallback to configured BASE_URL or APP_URL in .env
    base_url = os.environ.get('BASE_URL') or os.environ.get('APP_URL')
    if base_url:
        if not base_url.endswith('/'):
            base_url += '/'
        return base_url

    return 'http://localhost:5000/'

def _get_smtp_config():
    """Return SMTP config from .env, falling back to Gmail defaults."""
    host = os.environ.get('MAIL_SERVER', 'smtp.gmail.com')
    port = int(os.environ.get('MAIL_PORT', 587))
    user = os.environ.get('EMAIL_USER', 'info.loginpanel@gmail.com')
    password = os.environ.get('EMAIL_PASS', 'wedbfepklgtwtugf')
    use_ssl = os.environ.get('MAIL_USE_SSL', 'False').lower() in ('true', '1', 'yes')
    use_tls = os.environ.get('MAIL_USE_TLS', 'True').lower() in ('true', '1', 'yes')
    return host, port, user, password, use_ssl, use_tls

def _send_email(to_email, subject, html_content):
    try:
        host, port, user, password, use_ssl, use_tls = _get_smtp_config()

        msg = MIMEMultipart('alternative')
        msg['Subject'] = subject
        msg['From'] = user
        msg['To'] = to_email

        part = MIMEText(html_content, 'html')
        msg.attach(part)

        if use_ssl:
            server = smtplib.SMTP_SSL(host, port)
        else:
            server = smtplib.SMTP(host, port)
            if use_tls:
                server.starttls()

        server.login(user, password)
        server.send_message(msg)
        server.quit()
        return True
    except Exception as e:
        print(f"SMTP Error: {e}")
        return False

def send_otp_email(to_email, otp):
    html_content = f"""
    <html>
      <body style="font-family: Arial, sans-serif; background-color: #f0f3fa; padding: 20px; text-align: center;">
        <div style="max-width: 500px; margin: 0 auto; background: #ffffff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.05);">
          <h2 style="color: #0A2558; margin-bottom: 20px;">OoA Portal</h2>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 10px;">You requested to reset your security passcode.</p>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 25px;">Please use the following 4-digit OTP to proceed:</p>
          <div style="background: #f8f9fb; border: 2px dashed #0A2558; border-radius: 8px; padding: 15px; font-size: 32px; font-weight: bold; letter-spacing: 8px; color: #0A2558; margin-bottom: 30px;">
            {otp}
          </div>
          <p style="color: #8892aa; font-size: 12px;">If you did not request this, please ignore this email.</p>
        </div>
      </body>
    </html>
    """
    return _send_email(to_email, 'Your Security Passcode OTP', html_content)

def send_faculty_submission_email(hod_email, faculty_name, course_name, form_type, hod_name):
    html_content = f"""
    <html>
      <body style="font-family: Arial, sans-serif; background-color: #f0f3fa; padding: 20px; text-align: center;">
        <div style="max-width: 500px; margin: 0 auto; background: #ffffff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.05);">
          <h2 style="color: #0A2558; margin-bottom: 20px;">OoA Portal</h2>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 10px;">Dear {hod_name},</p>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 25px;">Faculty member <strong>{faculty_name}</strong> has submitted their {form_type} checklist for the course <strong>{course_name}</strong>.</p>
          <p style="color: #8892aa; font-size: 12px;">You can log in to the portal to review this submission.</p>
        </div>
      </body>
    </html>
    """
    return _send_email(hod_email, f'Faculty Submission Update: {course_name}', html_content)

def send_faculty_confirmation_email(faculty_email, faculty_name, course_name, form_type):
    html_content = f"""
    <html>
      <body style="font-family: Arial, sans-serif; background-color: #f0f3fa; padding: 20px; text-align: center;">
        <div style="max-width: 500px; margin: 0 auto; background: #ffffff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.05);">
          <h2 style="color: #0A2558; margin-bottom: 20px;">OoA Portal</h2>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 10px;">Dear {faculty_name},</p>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 25px;">Thank you for completing your checklist! Your submission for the course <strong>{course_name}</strong> ({form_type}) has been successfully logged.</p>
          <p style="color: #8892aa; font-size: 12px;">This is an automated confirmation of your submission.</p>
        </div>
      </body>
    </html>
    """
    return _send_email(faculty_email, f'Confirmation: Submission Successful for {course_name}', html_content)

def send_deadline_reminder_email(to_email, name, module_name, deadline_str):
    login_url = get_base_url() + "erp/login"
    html_content = f"""
    <html>
      <body style="font-family: Arial, sans-serif; background-color: #f0f3fa; padding: 20px; text-align: center;">
        <div style="max-width: 500px; margin: 0 auto; background: #ffffff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.05);">
          <h2 style="color: #0A2558; margin-bottom: 20px;">Semester Readiness & Closure Portal</h2>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 10px;">Dear HOD {name},</p>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 25px;">This is a friendly reminder that the submission window for <strong>{module_name}</strong> will close in <strong>2 days</strong>.</p>
          <p style="color: #ef4444; font-weight: bold; font-size: 16px; margin-bottom: 25px;">Official Deadline: {deadline_str.replace('T', ' ')}</p>
          <p style="color: #3d4460; font-size: 14px; margin-bottom: 25px;">Please log in to the portal as soon as possible to complete and finalize your checklist before access is closed.</p>
          <div style="margin-bottom: 30px;">
            <a href="{login_url}" style="background-color: #0A2558; color: #ffffff; padding: 12px 24px; border-radius: 6px; font-weight: bold; text-decoration: none; display: inline-block;">Go to Login Portal</a>
          </div>
          <p style="color: #8892aa; font-size: 12px;">This is an automated reminder. If you have already finalized your submission, please ignore this email.</p>
        </div>
      </body>
    </html>
    """
    return _send_email(to_email, f'IMPORTANT Reminder: {module_name} Submission Deadline is in 2 Days!', html_content)

def send_final_hour_deadline_email(to_email, name, module_name, deadline_str):
    login_url = get_base_url() + "erp/login"
    html_content = f"""
    <html>
      <body style="font-family: Arial, sans-serif; background-color: #fef2f2; padding: 20px; text-align: center;">
        <div style="max-width: 500px; margin: 0 auto; background: #ffffff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.05); border-top: 6px solid #ef4444;">
          <h2 style="color: #b91c1c; margin-bottom: 20px;">⚠️ FINAL WARNING</h2>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 10px;">Dear HOD {name},</p>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 25px;">This is a critical reminder that the submission window for <strong>{module_name}</strong> will close in <strong>exactly 1 hour</strong>!</p>
          <p style="color: #ef4444; font-weight: bold; font-size: 18px; margin-bottom: 25px; background: #fee2e2; padding: 10px; border-radius: 6px; display: inline-block;">
            CLOSING DEADLINE: {deadline_str.replace('T', ' ')}
          </p>
          <p style="color: #3d4460; font-size: 14px; margin-bottom: 25px;">Late submissions will be automatically blocked. Please finalize and submit your form immediately to avoid lockout.</p>
          <div style="margin-bottom: 30px;">
            <a href="{login_url}" style="background-color: #ef4444; color: #ffffff; padding: 12px 24px; border-radius: 6px; font-weight: bold; text-decoration: none; display: inline-block;">Submit Form Now</a>
          </div>
          <p style="color: #8892aa; font-size: 12px;">This is an automated critical alert. If you have already finalized your submission, please ignore this email.</p>
        </div>
      </body>
    </html>
    """
    return _send_email(to_email, f'⚠️ CRITICAL: {module_name} Submission Deadline is in 1 HOUR!', html_content)

def check_and_send_deadline_reminders():
    settings = get_global_settings()
    now = datetime.now()
    
    # Check Readiness
    r_deadline = settings.get('readiness_deadline')
    if r_deadline:
        try:
            r_dt = datetime.strptime(r_deadline, "%Y-%m-%dT%H:%M")
            diff_hours = (r_dt - now).total_seconds() / 3600.0
            r_date_str = r_dt.strftime("%d %B %Y")
            
            # Send 2-day warning
            if 0 < diff_hours <= 48:
                for user in users_col.find():
                    email = user.get('email')
                    name = user.get('name')
                    has_sub = submissions_col.find_one({
                        'identity.submitterEmail': email,
                        'form_type': 'readiness',
                        '_draft': False
                    })
                    if not has_sub:
                        reminder_key = f"reminder_readiness_2day_{r_deadline}_{email}"
                        if not db['sent_reminders'].find_one({'_id': reminder_key}):
                            print(f"Sending 2-day deadline warning to HOD {name} ({email}) for Readiness...")
                            send_deadline_reminder_email(email, name, "Semester Readiness", r_date_str)
                            db['sent_reminders'].insert_one({'_id': reminder_key, 'sent_at': datetime.utcnow().isoformat()})
            
            # Send 1-hour warning
            if 0 < diff_hours <= 1.0:
                for user in users_col.find():
                    email = user.get('email')
                    name = user.get('name')
                    has_sub = submissions_col.find_one({
                        'identity.submitterEmail': email,
                        'form_type': 'readiness',
                        '_draft': False
                    })
                    if not has_sub:
                        reminder_key = f"reminder_readiness_1hour_{r_deadline}_{email}"
                        if not db['sent_reminders'].find_one({'_id': reminder_key}):
                            print(f"Sending 1-hour deadline warning to HOD {name} ({email}) for Readiness...")
                            send_final_hour_deadline_email(email, name, "Semester Readiness", r_date_str)
                            db['sent_reminders'].insert_one({'_id': reminder_key, 'sent_at': datetime.utcnow().isoformat()})
                            
        except Exception as e:
            print(f"Error checking readiness deadline reminders: {e}")

    # Check Closure
    c_deadline = settings.get('closure_deadline')
    if c_deadline:
        try:
            c_dt = datetime.strptime(c_deadline, "%Y-%m-%dT%H:%M")
            diff_hours = (c_dt - now).total_seconds() / 3600.0
            c_date_str = c_dt.strftime("%d %B %Y")
            
            # Send 2-day warning
            if 0 < diff_hours <= 48:
                for user in users_col.find():
                    email = user.get('email')
                    name = user.get('name')
                    has_sub = submissions_col.find_one({
                        'identity.submitterEmail': email,
                        'form_type': 'closure',
                        '_draft': False
                    })
                    if not has_sub:
                        reminder_key = f"reminder_closure_2day_{c_deadline}_{email}"
                        if not db['sent_reminders'].find_one({'_id': reminder_key}):
                            print(f"Sending 2-day deadline warning to HOD {name} ({email}) for Closure...")
                            send_deadline_reminder_email(email, name, "Semester Closure", c_date_str)
                            db['sent_reminders'].insert_one({'_id': reminder_key, 'sent_at': datetime.utcnow().isoformat()})
            
            # Send 1-hour warning
            if 0 < diff_hours <= 1.0:
                for user in users_col.find():
                    email = user.get('email')
                    name = user.get('name')
                    has_sub = submissions_col.find_one({
                        'identity.submitterEmail': email,
                        'form_type': 'closure',
                        '_draft': False
                    })
                    if not has_sub:
                        reminder_key = f"reminder_closure_1hour_{c_deadline}_{email}"
                        if not db['sent_reminders'].find_one({'_id': reminder_key}):
                            print(f"Sending 1-hour deadline warning to HOD {name} ({email}) for Closure...")
                            send_final_hour_deadline_email(email, name, "Semester Closure", c_date_str)
                            db['sent_reminders'].insert_one({'_id': reminder_key, 'sent_at': datetime.utcnow().isoformat()})
                            
        except Exception as e:
            print(f"Error checking closure deadline reminders: {e}")

def start_deadline_scheduler():
    import threading
    import time
    def run_scheduler():
        print("Deadline reminder background scheduler started...")
        while True:
            try:
                check_and_send_deadline_reminders()
            except Exception as ex:
                print(f"Scheduler check error: {ex}")
            time.sleep(60)  # check every 60 seconds

    t = threading.Thread(target=run_scheduler, daemon=True)
    t.start()

    t = threading.Thread(target=run_scheduler, daemon=True)
    t.start()

# ── Helpers ───────────────────────────────────────────────

def is_deadline_passed(deadline_str):
    if not deadline_str:
        return False
    try:
        from datetime import datetime
        dt = datetime.strptime(deadline_str, "%Y-%m-%dT%H:%M")
        return datetime.now() > dt
    except Exception as e:
        print(f"Error parsing deadline {deadline_str}: {e}")
        return False
def get_global_settings():
    settings = settings_col.find_one({'_id': 'global'}) or {}
    settings.setdefault('readiness_enabled', True)
    settings.setdefault('readiness_deadline', '')
    settings.setdefault('closure_enabled', True)
    settings.setdefault('closure_deadline', '')
    settings.setdefault('enabled_years', ['2024-25', '2025-26', '2026-27', '2027-28'])
    settings.setdefault('enabled_semesters', ['Even', 'Odd'])
    return settings
def _sections_for(form_type):
    """Return (hod_sections_or_report, faculty_sections) for the given form type."""
    if form_type == 'closure':
        return HOD_CLOSURE_SECTIONS, FACULTY_CLOSURE_SECTIONS
    return HOD_SECTIONS, FACULTY_SECTIONS

# ── Routes ──────────────────────────────────────────────

@erp_bp.route('/')
def index():
    if 'user_email' not in session:
        return redirect(url_for('erp.login'))
    user_doc = users_col.find_one({'email': session['user_email']})
    user = {
        'email': session['user_email'],
        'name': session['user_name'],
        'department': user_doc.get('department', '') if user_doc else '',
        'timeout_pref': user_doc.get('timeout_pref', 15) if user_doc else 15,
        'lock_enabled': user_doc.get('lock_enabled', False) if user_doc else False,
        'first_time_login': user_doc.get('first_time_login', True) if user_doc else False
    }
    settings = settings_col.find_one({'_id': 'global'}) or {
        'readiness_enabled': True, 'readiness_deadline': '',
        'closure_enabled': True, 'closure_deadline': ''
    }
    return render_template('erp/dashboard.html', user=user, settings=settings, departments=DEPARTMENTS)

def has_access_override(email, module_type):
    user = users_col.find_one({'email': email})
    if user:
        if module_type == 'readiness':
            return user.get('readiness_access_override', False)
        elif module_type == 'closure':
            return user.get('closure_access_override', False)
    return False

@erp_bp.route('/readiness')
def readiness():
    if 'user_email' not in session:
        return redirect(url_for('erp.login'))
    settings = get_global_settings()
    is_passed = is_deadline_passed(settings.get('readiness_deadline'))
    override = has_access_override(session['user_email'], 'readiness')
    req = db['access_requests'].find_one({'user_email': session['user_email'], 'module': 'Semester Readiness'})
    user = {'email': session['user_email'], 'name': session['user_name']}
    return render_template('erp/form.html', 
                           hod_sections=HOD_SECTIONS, 
                           departments=DEPARTMENTS, 
                           user=user, 
                           view_mode='list', 
                           settings=settings,
                           is_deadline_passed=is_passed,
                           has_override=override,
                           extension_request=req)

@erp_bp.route('/readiness/form')
@erp_bp.route('/readiness/form/<sub_id>')
def readiness_form(sub_id=None):
    if 'user_email' not in session:
        return redirect(url_for('erp.login'))
    settings = get_global_settings()
    is_passed = is_deadline_passed(settings.get('readiness_deadline'))
    override = has_access_override(session['user_email'], 'readiness')
    if not session.get('admin') and is_passed and not override:
        return render_template('erp/deadline_passed.html', module='Semester Readiness', deadline=settings.get('readiness_deadline'))
    user = {'email': session['user_email'], 'name': session['user_name']}
    return render_template('erp/form.html', hod_sections=HOD_SECTIONS, departments=DEPARTMENTS, user=user, view_mode='form', edit_id=sub_id, settings=settings)

@erp_bp.route('/closure')
def closure():
    if 'user_email' not in session:
        return redirect(url_for('erp.login'))
    settings = get_global_settings()
    is_passed = is_deadline_passed(settings.get('closure_deadline'))
    override = has_access_override(session['user_email'], 'closure')
    req = db['access_requests'].find_one({'user_email': session['user_email'], 'module': 'Semester Closure'})
    user = {'email': session['user_email'], 'name': session['user_name']}
    return render_template('erp/closure.html',
                           hod_closure_sections=HOD_CLOSURE_SECTIONS,
                           departments=DEPARTMENTS,
                           user=user,
                           view_mode='list',
                           settings=settings,
                           is_deadline_passed=is_passed,
                           has_override=override,
                           extension_request=req)

@erp_bp.route('/closure/form')
@erp_bp.route('/closure/form/<sub_id>')
def closure_form(sub_id=None):
    if 'user_email' not in session:
        return redirect(url_for('erp.login'))
    settings = get_global_settings()
    is_passed = is_deadline_passed(settings.get('closure_deadline'))
    override = has_access_override(session['user_email'], 'closure')
    if not session.get('admin') and is_passed and not override:
        return render_template('erp/deadline_passed.html', module='Semester Closure', deadline=settings.get('closure_deadline'))
    user = {'email': session['user_email'], 'name': session['user_name']}
    return render_template('erp/closure.html',
                           hod_closure_sections=HOD_CLOSURE_SECTIONS,
                           departments=DEPARTMENTS,
                           user=user,
                           view_mode='form',
                           edit_id=sub_id,
                           settings=settings)

@erp_bp.route('/login')
def login():
    # Separate ERP login removed — SSO via the Jain Hub session.
    if session.get('user_email'):
        return redirect(url_for('erp.index'))
    return redirect('/login')

# ═════════════════════════════════════════════════════════════
# FACULTY SHARE LINK ROUTES — handles BOTH readiness AND closure
# ═════════════════════════════════════════════════════════════

@erp_bp.route('/faculty-form/<share_token>')
def faculty_form(share_token):
    """Public route where faculty fills their checklist using HOD's share link.
    Auto-detects whether the parent HOD submission is for readiness or closure
    and renders the appropriate checklist sections.
    """
    hod_submission = submissions_col.find_one({'share_token': share_token})
    if not hod_submission:
        return render_template('erp/faculty_link_invalid.html'), 404

    form_type = hod_submission.get('form_type', 'readiness')
    _, faculty_sections = _sections_for(form_type)

    # Auto-fill context from HOD submission so faculty doesn't fill it again
    context = {
        'share_token': share_token,
        'form_type': form_type,
        'hod_name': hod_submission.get('identity', {}).get('hodName', '')
                     or hod_submission.get('identity', {}).get('hod', ''),
        'hod_email': hod_submission.get('identity', {}).get('hodEmail', '')
                      or hod_submission.get('identity', {}).get('email', ''),
        'dept': hod_submission.get('identity', {}).get('dept', ''),
        'campus': hod_submission.get('identity', {}).get('campus', ''),
        'semester': hod_submission.get('identity', {}).get('semester', '')
                     or hod_submission.get('identity', {}).get('sem', ''),
        'ac_year': hod_submission.get('identity', {}).get('acYear', '')
                    or hod_submission.get('identity', {}).get('ac_year', ''),
        'programs': hod_submission.get('identity', {}).get('programs', ''),
        'submission_id': str(hod_submission['_id']),
        'faculty_sections': faculty_sections,
    }
    return render_template('erp/faculty_form.html', **context)

@erp_bp.route('/api/faculty-submit', methods=['POST'])
def faculty_submit():
    """Faculty submits their checklist (no login required, uses share_token)."""
    data = request.json
    share_token = data.get('share_token')

    if not share_token:
        return jsonify({'ok': False, 'error': 'Invalid share link'})

    hod_submission = submissions_col.find_one({'share_token': share_token})
    if not hod_submission:
        return jsonify({'ok': False, 'error': 'Share link not found'})

    form_type = hod_submission.get('form_type', 'readiness')

    # Build faculty submission doc
    faculty_doc = {
        'parent_submission_id': str(hod_submission['_id']),
        'share_token': share_token,
        'form_type': form_type,  # tag with type so we can render right view
        'dept': hod_submission.get('identity', {}).get('dept', ''),
        'campus': hod_submission.get('identity', {}).get('campus', ''),
        'semester': hod_submission.get('identity', {}).get('semester', '')
                     or hod_submission.get('identity', {}).get('sem', ''),
        'ac_year': hod_submission.get('identity', {}).get('acYear', ''),
        'hod_name': hod_submission.get('identity', {}).get('hodName', '')
                     or hod_submission.get('identity', {}).get('hod', ''),
        'hod_email': hod_submission.get('identity', {}).get('hodEmail', '')
                      or hod_submission.get('identity', {}).get('email', ''),
        'faculty_name':   data.get('faculty_name', '').strip(),
        'faculty_email':  data.get('faculty_email', '').strip(),
        'course_name':    data.get('course_name', '').strip(),
        'course_code':    data.get('course_code', '').strip(),
        'program':        data.get('program', '').strip(),
        'year_sem':       data.get('year_sem', '').strip(),
        'no_of_students': data.get('no_of_students', '').strip(),
        'checklist':      data.get('checklist', {}),
        'hod_remarks': '',
        'hod_review_status': 'pending',
        'timestamp': datetime.utcnow().isoformat(),
    }

    # Same faculty+course already submitted? -> update, else insert
    existing = faculty_submissions_col.find_one({
        'parent_submission_id': str(hod_submission['_id']),
        'faculty_email': faculty_doc['faculty_email'],
        'course_code': faculty_doc['course_code']
    })

    if existing:
        faculty_submissions_col.update_one(
            {'_id': existing['_id']},
            {'$set': faculty_doc}
        )
    else:
        result = faculty_submissions_col.insert_one(faculty_doc)

    # Send email to HOD and Faculty
    hod_email = faculty_doc['hod_email']
    hod_name = faculty_doc['hod_name']
    if hod_email:
        send_faculty_submission_email(
            hod_email=hod_email, 
            faculty_name=faculty_doc['faculty_name'], 
            course_name=faculty_doc['course_name'], 
            form_type=faculty_doc['form_type'],
            hod_name=hod_name
        )
    if faculty_doc['faculty_email']:
        send_faculty_confirmation_email(
            faculty_email=faculty_doc['faculty_email'],
            faculty_name=faculty_doc['faculty_name'],
            course_name=faculty_doc['course_name'],
            form_type=faculty_doc['form_type']
        )

    if existing:
        return jsonify({'ok': True, 'updated': True, 'id': str(existing['_id'])})
    else:
        return jsonify({'ok': True, 'updated': False, 'id': str(result.inserted_id)})

@erp_bp.route('/api/faculty-submissions/<sub_id>')
def get_faculty_submissions(sub_id):
    """HOD views all faculty submissions for their form."""
    if 'user_email' not in session and not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Not logged in'})

    docs = list(faculty_submissions_col.find({'parent_submission_id': sub_id}, sort=[('timestamp', -1)]))
    for d in docs:
        d['_id'] = str(d['_id'])
    return jsonify({'ok': True, 'submissions': docs})

@erp_bp.route('/api/faculty-review/<faculty_sub_id>', methods=['POST'])
def review_faculty_submission(faculty_sub_id):
    """HOD reviews and adds remarks/status to a faculty submission."""
    if 'user_email' not in session:
        return jsonify({'ok': False, 'error': 'Not logged in'})

    data = request.json
    update = {
        'hod_remarks': data.get('hod_remarks', ''),
        'hod_review_status': data.get('hod_review_status', 'reviewed'),
        'reviewed_at': datetime.utcnow().isoformat(),
    }
    faculty_submissions_col.update_one({'_id': ObjectId(faculty_sub_id)}, {'$set': update})
    return jsonify({'ok': True})

@erp_bp.route('/api/faculty-delete/<faculty_sub_id>', methods=['POST'])
def delete_faculty_submission(faculty_sub_id):
    """HOD or Admin deletes a faculty submission."""
    if 'user_email' not in session and not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Not logged in'})
    faculty_submissions_col.delete_one({'_id': ObjectId(faculty_sub_id)})
    return jsonify({'ok': True})

# ═════════════════════════════════════════════════════════════
# HOD SUBMISSION (generates share_token used by faculty link)
# ═════════════════════════════════════════════════════════════

@erp_bp.route('/api/submit', methods=['POST'])
def submit():
    data = request.json
    sub_id = data.get('_id')

    # Security check for late submission
    form_type = data.get('form_type', 'readiness')
    settings = get_global_settings()
    deadline = settings.get('readiness_deadline') if form_type == 'readiness' else settings.get('closure_deadline')
    if not session.get('admin') and is_deadline_passed(deadline):
        if not has_access_override(session.get('user_email'), form_type):
            return jsonify({'ok': False, 'error': f'The deadline has passed. Late submissions are blocked. Please request admin access.'})

    if '_id' in data:
        del data['_id']

    data['timestamp'] = datetime.utcnow().isoformat()
    is_draft = data.get('_draft', False)

    if sub_id:
        existing = submissions_col.find_one({'_id': ObjectId(sub_id)})
        if not existing:
            return jsonify({'ok': False, 'error': 'Submission not found'})

        edit_count = existing.get('edit_count', 0)
        edit_request = existing.get('edit_request', {})
        
        # We only count it as an edit if the previous save was NOT a draft
        was_previously_submitted = not existing.get('_draft', False)

        if not is_draft and was_previously_submitted:
            if edit_count >= 2 and edit_request.get('status') != 'approved':
                return jsonify({'ok': False, 'error': 'Edit limit exceeded. Please request admin approval.'})

            if edit_request.get('status') == 'approved':
                data['edit_request'] = None

            data['edit_count'] = edit_count + 1
        else:
            # Saving a draft, or finalizing a draft for the first time keeps current edit_count
            data['edit_count'] = edit_count

        # Preserve share_token if exists
        if existing.get('share_token'):
            data['share_token'] = existing['share_token']
        else:
            data['share_token'] = secrets.token_urlsafe(16)

        submissions_col.update_one({'_id': ObjectId(sub_id)}, {'$set': data})
        return jsonify({'ok': True, 'id': sub_id, 'share_token': data['share_token']})
    else:
        data['edit_count'] = 0
        data['share_token'] = secrets.token_urlsafe(16)
        result = submissions_col.insert_one(data)
        return jsonify({'ok': True, 'id': str(result.inserted_id), 'share_token': data['share_token']})

@erp_bp.route('/api/upload-evidence', methods=['POST'])
def upload_evidence():
    if 'file' not in request.files:
        return jsonify({'ok': False, 'error': 'No file part'})
    file = request.files['file']
    if file.filename == '':
        return jsonify({'ok': False, 'error': 'No selected file'})

    if file:
        filename = secure_filename(file.filename)
        ts = datetime.utcnow().strftime("%Y%m%d%H%M%S")
        filename = f"{ts}_{filename}"

        upload_path = os.path.join('static', 'uploads')
        os.makedirs(upload_path, exist_ok=True)

        file.save(os.path.join(upload_path, filename))
        url = url_for('static', filename=f"uploads/{filename}")
        return jsonify({'ok': True, 'url': url})

    return jsonify({'ok': False, 'error': 'Unknown error'})

@erp_bp.route('/api/request-edit/<sub_id>', methods=['POST'])
def request_edit(sub_id):
    data = request.json
    comment = data.get('comment', '').strip()

    submissions_col.update_one(
        {'_id': ObjectId(sub_id)},
        {'$set': {'edit_request': {'pending': True, 'comment': comment, 'status': 'pending', 'timestamp': datetime.utcnow().isoformat()}}}
    )
    return jsonify({'ok': True})

@erp_bp.route('/api/register', methods=['POST'])
def register():
    data = request.json
    name = data.get('name', '').strip()
    email = data.get('email', '').strip()
    password = data.get('password', '')

    if not name or not email or not password:
        return jsonify({'ok': False, 'error': 'Name, Email, and Password are required'})

    existing = users_col.find_one({'email': email})
    if existing:
        return jsonify({'ok': False, 'error': 'You already have an account. Please use Login.'})

    users_col.insert_one({
        'name': name,
        'email': email,
        'password': generate_password_hash(password),
        'created_at': datetime.utcnow().isoformat(),
        'first_time_login': True
    })
    session['user_email'] = email
    session['user_name'] = name
    return jsonify({'ok': True})

@erp_bp.route('/api/login', methods=['POST'])
def api_login():
    data = request.json
    email = data.get('email', '').strip()
    password = data.get('password', '')

    if not email or not password:
        return jsonify({'ok': False, 'error': 'Email and Password are required'})

    user = users_col.find_one({'email': email})
    if not user:
        return jsonify({'ok': False, 'error': 'User not found. Please register as a New User.'})

    if not user.get('password'):
        return jsonify({'ok': False, 'error': 'Please use Forgot Password to set your initial password.'})

    if not check_password_hash(user['password'], password):
        return jsonify({'ok': False, 'error': 'Invalid credentials'})

    session['user_email'] = user['email']
    session['user_name'] = user['name']
    return jsonify({'ok': True})

# Removed passcode endpoints

@erp_bp.route('/api/forgot-password', methods=['POST'])
def forgot_password():
    email = request.json.get('email', '').strip()
    user = users_col.find_one({'email': email})
    if not user:
        return jsonify({'ok': False, 'error': 'User not found'})

    otp = str(random.randint(1000, 9999))
    users_col.update_one({'email': email}, {'$set': {'reset_otp': otp}})

    send_otp_email(email, otp)

    session['pending_email'] = email
    return jsonify({'ok': True})

@erp_bp.route('/api/reset-password', methods=['POST'])
def reset_password():
    email = session.get('pending_email') or session.get('user_email')
    if not email:
        return jsonify({'ok': False, 'error': 'Session expired. Please try again.'})
    data = request.json
    otp = data.get('otp')
    new_password = data.get('password')

    if not otp or not new_password or len(str(new_password)) < 6:
        return jsonify({'ok': False, 'error': 'Invalid OTP or password must be at least 6 characters'})

    user = users_col.find_one({'email': email})
    if str(user.get('reset_otp')) != str(otp):
        return jsonify({'ok': False, 'error': 'Invalid OTP'})

    users_col.update_one({'email': email}, {'$set': {'password': generate_password_hash(new_password), 'reset_otp': None}})
    session['user_email'] = email
    session['user_name'] = user['name']
    session.pop('pending_email', None)
    return jsonify({'ok': True})

@erp_bp.route('/api/update-profile', methods=['POST'])
def update_profile():
    if 'user_email' not in session:
        return jsonify({'ok': False, 'error': 'Not logged in'})
    dept = request.json.get('department')
    users_col.update_one({'email': session['user_email']}, {'$set': {
        'department': dept,
        'first_time_login': False
    }})
    return jsonify({'ok': True})

@erp_bp.route('/api/logout', methods=['POST', 'GET'])
def logout():
    # SSO: leaving the ERP returns you to the Jain Hub (one session everywhere)
    if request.method == 'GET':
        return redirect('/home')
    return jsonify({'ok': True, 'redirect': '/home'})

@erp_bp.route('/api/my-submissions')
def my_submissions():
    if 'user_email' not in session:
        return jsonify({'ok': False, 'error': 'Not logged in'})

    email = session['user_email']
    form_type = request.args.get('type')

    query = {'identity.submitterEmail': email}
    if form_type:
        query['form_type'] = form_type

    docs = list(submissions_col.find(query, sort=[('timestamp', -1)]))
    for d in docs:
        d['_id'] = str(d['_id'])
        # Attach faculty submission count
        d['faculty_submission_count'] = faculty_submissions_col.count_documents({'parent_submission_id': d['_id']})
    return jsonify({'ok': True, 'submissions': docs})

@erp_bp.route('/api/get-submission/<sub_id>')
def get_submission(sub_id):
    """Get a single HOD submission with its faculty submissions (for share/review page)."""
    if 'user_email' not in session and not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Not logged in'})
    try:
        doc = submissions_col.find_one({'_id': ObjectId(sub_id)})
        if not doc:
            return jsonify({'ok': False, 'error': 'Not found'})
        doc['_id'] = str(doc['_id'])
        fac_docs = list(faculty_submissions_col.find({'parent_submission_id': sub_id}, sort=[('timestamp', -1)]))
        for f in fac_docs:
            f['_id'] = str(f['_id'])
        doc['_faculty_submissions'] = fac_docs
        return jsonify({'ok': True, 'submission': doc})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})

@erp_bp.route('/api/export/<submission_id>')
def export_submission(submission_id):
    """Export the full submission (HOD + all faculty) as Excel — admin/HOD download."""
    try:
        doc = submissions_col.find_one({'_id': ObjectId(submission_id)})
    except:
        return "Not found", 404
    if not doc:
        return "Not found", 404
    fac_docs = list(faculty_submissions_col.find({'parent_submission_id': submission_id}))
    doc['_faculty_submissions'] = fac_docs
    wb = build_workbook([doc])
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    dept = doc.get('identity', {}).get('dept', 'Department').replace(' ', '_')[:30]
    form_type = doc.get('form_type', 'readiness').title()
    return send_file(buf, as_attachment=True,
                     download_name=f"Sem{form_type}_{dept}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@erp_bp.route('/api/export-hod-checklist/<submission_id>')
def export_hod_checklist(submission_id):
    """Download the HOD's filled checklist as an Excel template
    that the HOD can hand to faculty as a reference."""
    try:
        doc = submissions_col.find_one({'_id': ObjectId(submission_id)})
    except:
        return "Not found", 404
    if not doc:
        return "Not found", 404

    form_type = doc.get('form_type', 'readiness')
    wb = build_hod_checklist_only(doc, form_type)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    dept = doc.get('identity', {}).get('dept', 'Department').replace(' ', '_')[:30]
    label = 'Closure' if form_type == 'closure' else 'Readiness'
    return send_file(buf, as_attachment=True,
                     download_name=f"HOD_{label}_{dept}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

# ── Admin ────────────────────────────────────────────────

@erp_bp.route('/admin/login', methods=['GET', 'POST'])
def admin_login():
    # Hub admins are auto-admitted via SSO; the form remains as a fallback.
    if request.method == 'GET' and session.get('admin'):
        return redirect(url_for('erp.admin_dashboard'))
    if request.method == 'POST':
        u = request.form.get('username')
        p = request.form.get('password')
        if u == ADMIN_USERNAME and p == ADMIN_PASSWORD:
            session['admin'] = True
            return redirect(url_for('erp.admin_dashboard'))
        flash('Invalid credentials')
    return render_template('erp/admin_login.html')

@erp_bp.route('/admin/logout')
def admin_logout():
    session.pop('admin', None)
    return redirect(url_for('erp.admin_login'))

@erp_bp.route('/admin/impersonate/<email>')
def admin_impersonate(email):
    if not session.get('admin'):
        return redirect(url_for('erp.admin_login'))
    user = users_col.find_one({'email': email})
    if not user:
        return "User not found", 404
    session['user_email'] = user['email']
    session['user_name'] = user['name']
    return redirect('/erp/')

@erp_bp.route('/admin')
def admin_dashboard():
    if not session.get('admin'):
        return redirect(url_for('erp.admin_login'))
    submissions = list(submissions_col.find().sort('timestamp', -1))
    for s in submissions:
        s['_id'] = str(s['_id'])
        s['faculty_submission_count'] = faculty_submissions_col.count_documents({'parent_submission_id': s['_id']})

    users = list(users_col.find().sort('created_at', -1))
    for u in users:
        u['_id'] = str(u['_id'])
        latest_sub = submissions_col.find_one({'identity.submitterEmail': u['email']}, sort=[('timestamp', -1)])
        u['latest_dept'] = u.get('department') or (latest_sub['identity']['dept'] if latest_sub and 'identity' in latest_sub else 'N/A')

    access_requests = list(db['access_requests'].find().sort('timestamp', -1))
    for r in access_requests:
        r['_id'] = str(r['_id'])

    settings = settings_col.find_one({'_id': 'global'}) or {}
    settings.setdefault('readiness_enabled', True)
    settings.setdefault('readiness_deadline', '')
    settings.setdefault('closure_enabled', True)
    settings.setdefault('closure_deadline', '')
    settings.setdefault('enabled_years', ['2024-25', '2025-26', '2026-27', '2027-28'])
    settings.setdefault('enabled_semesters', ['Even', 'Odd'])
    
    pwa_stats = db['pwa_analytics'].find_one({'_id': 'pwa_stats'}) or {'installs': 0, 'launches': 0}
    
    notifications = list(db['notifications'].find().sort('created_at', -1).limit(10))
    for n in notifications:
        n['_id'] = str(n['_id'])
    
    return render_template('erp/admin.html',
                           submissions=submissions,
                           hod_sections=HOD_SECTIONS,
                           faculty_sections=FACULTY_SECTIONS,
                           users=users, settings=settings, departments=DEPARTMENTS,
                           access_requests=access_requests,
                           pwa_stats=pwa_stats,
                           notifications=notifications)

@erp_bp.route('/admin/settings', methods=['POST'])
def save_settings():
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    data = request.json
    
    prev_settings = settings_col.find_one({'_id': 'global'}) or {}
    
    settings_col.update_one(
        {'_id': 'global'},
        {'$set': {
            'readiness_enabled': data.get('readiness_enabled', True),
            'readiness_deadline': data.get('readiness_deadline', ''),
            'closure_enabled': data.get('closure_enabled', True),
            'closure_deadline': data.get('closure_deadline', ''),
            'enabled_years': data.get('enabled_years', ['2024-25', '2025-26', '2026-27', '2027-28']),
            'enabled_semesters': data.get('enabled_semesters', ['Even', 'Odd'])
        }},
        upsert=True
    )
    
    was_closure_enabled = prev_settings.get('closure_enabled', False)
    is_closure_enabled = data.get('closure_enabled', True)
    
    prev_readiness_dl = prev_settings.get('readiness_deadline', '')
    new_readiness_dl = data.get('readiness_deadline', '')
    prev_closure_dl = prev_settings.get('closure_deadline', '')
    new_closure_dl = data.get('closure_deadline', '')
    
    notif_body = ""
    notif_title = ""
    
    if not was_closure_enabled and is_closure_enabled:
        notif_title = "Semester Closure Added"
        notif_body = "Semester Closure has been added! Please check your portal."
    elif new_closure_dl != prev_closure_dl and new_closure_dl:
        notif_title = "Closure Deadline Extended"
        notif_body = f"Semester Closure last date has been updated/extended to {new_closure_dl}!"
    elif new_readiness_dl != prev_readiness_dl and new_readiness_dl:
        notif_title = "Readiness Deadline Extended"
        notif_body = f"Semester Readiness last date has been updated/extended to {new_readiness_dl}!"
        
    if notif_title:
        db['notifications'].insert_one({
            'title': notif_title,
            'body': notif_body,
            'created_at': datetime.utcnow().isoformat(),
            'type': 'broadcast'
        })
        
    return jsonify({'ok': True})

@erp_bp.route('/api/get-notifications', methods=['GET'])
def get_notifications():
    notifs = list(db['notifications'].find().sort('created_at', -1).limit(10))
    for n in notifs:
        n['_id'] = str(n['_id'])
    return jsonify({'ok': True, 'notifications': notifs})

@erp_bp.route('/api/send-notification', methods=['POST'])
def send_custom_notification():
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    data = request.json or {}
    title = data.get('title', '').strip()
    body = data.get('body', '').strip()
    if not title or not body:
        return jsonify({'ok': False, 'error': 'Missing title or body'})
    
    db['notifications'].insert_one({
        'title': title,
        'body': body,
        'created_at': datetime.utcnow().isoformat(),
        'type': 'custom'
    })
    return jsonify({'ok': True})

@erp_bp.route('/admin/delete-notification/<notif_id>', methods=['POST'])
def delete_notification(notif_id):
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    from bson import ObjectId
    db['notifications'].delete_one({'_id': ObjectId(notif_id)})
    return jsonify({'ok': True})

@erp_bp.route('/api/request-extension', methods=['POST'])
def request_extension():
    if 'user_email' not in session:
        return jsonify({'ok': False, 'error': 'Not logged in'})
    data = request.json or {}
    comment = data.get('comment', '').strip()
    module = data.get('module', '')
    
    db['access_requests'].update_one(
        {
            'user_email': session['user_email'],
            'module': module
        },
        {
            '$set': {
                'user_name': session['user_name'],
                'comment': comment,
                'status': 'pending',
                'timestamp': datetime.utcnow().isoformat()
            }
        },
        upsert=True
    )
    return jsonify({'ok': True})

def send_override_approval_email(to_email, name, module_name, admin_comment):
    login_url = get_base_url() + "erp/login"
    html_content = f"""
    <html>
      <body style="font-family: Arial, sans-serif; background-color: #ecfdf5; padding: 20px; text-align: center;">
        <div style="max-width: 500px; margin: 0 auto; background: #ffffff; padding: 30px; border-radius: 12px; box-shadow: 0 4px 15px rgba(0,0,0,0.05); border-top: 6px solid #10b981;">
          <h2 style="color: #065f46; margin-bottom: 20px;">✅ Late Submission Access Approved</h2>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 10px;">Dear HOD {name},</p>
          <p style="color: #3d4460; font-size: 16px; margin-bottom: 25px;">The Office of Academics administration has reviewed and <strong>approved</strong> your request for late submission access to <strong>{module_name}</strong>.</p>

          <div style="margin-bottom: 25px; background: #f0fdf4; border-left: 4px solid #10b981; padding: 12px 15px; border-radius: 6px; text-align: left;">
            <strong style="color: #065f46; font-size: 13px; display: block; margin-bottom: 6px;">Administrator Remarks:</strong>
            <span style="color: #374151; font-size: 14px; font-style: italic;">"{admin_comment or 'No comment provided.'}"</span>
          </div>

          <p style="color: #3d4460; font-size: 14px; margin-bottom: 25px;">You may now log in and complete your submission form.</p>

          <div style="margin-bottom: 30px;">
            <a href="{login_url}" style="background-color: #10b981; color: #ffffff; padding: 12px 24px; border-radius: 6px; font-weight: bold; text-decoration: none; display: inline-block;">Go to Submission Portal</a>
          </div>

          <p style="color: #8892aa; font-size: 12px;">This is an automated notification. Please do not reply directly to this email.</p>
        </div>
      </body>
    </html>
    """
    return _send_email(to_email, f'✅ APPROVED: Late Submission Access Granted for {module_name}', html_content)

@erp_bp.route('/admin/approve-access/<request_id>', methods=['POST'])
def approve_access(request_id):
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    req = db['access_requests'].find_one({'_id': ObjectId(request_id)})
    if not req:
        return jsonify({'ok': False, 'error': 'Request not found'})
    
    data = request.json or {}
    admin_comment = data.get('admin_comment', '').strip()
    
    email = req['user_email']
    module = req['module']
    field = 'readiness_access_override' if 'Readiness' in module else 'closure_access_override'
    
    users_col.update_one({'email': email}, {'$set': {field: True}})
    db['access_requests'].update_one(
        {'_id': ObjectId(request_id)}, 
        {'$set': {'status': 'approved', 'admin_comment': admin_comment}}
    )
    
    # Send email notification to HOD
    send_override_approval_email(email, req.get('user_name', 'HOD'), module, admin_comment)
    
    return jsonify({'ok': True})

@erp_bp.route('/admin/revoke-access/<email>/<module_type>', methods=['POST'])
def revoke_access(email, module_type):
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    field = 'readiness_access_override' if module_type == 'readiness' else 'closure_access_override'
    users_col.update_one({'email': email}, {'$set': {field: False}})
    m_name = 'Semester Readiness' if module_type == 'readiness' else 'Semester Closure'
    db['access_requests'].delete_many({'user_email': email, 'module': m_name})
    return jsonify({'ok': True})

@erp_bp.route('/admin/send-manual-reminders', methods=['POST'])
def send_manual_reminders():
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    
    settings = get_global_settings()
    sent_count = 0
    
    # Readiness manual send
    r_deadline = settings.get('readiness_deadline')
    r_date_str = "N/A"
    if r_deadline:
        try:
            r_dt = datetime.strptime(r_deadline, "%Y-%m-%dT%H:%M")
            r_date_str = r_dt.strftime("%d %B %Y")
        except Exception as ex:
            print(f"Error parsing manual readiness deadline: {ex}")
        
        for user in users_col.find():
            email = user.get('email')
            name = user.get('name')
            has_sub = submissions_col.find_one({
                'identity.submitterEmail': email,
                'form_type': 'readiness',
                '_draft': False
            })
            if not has_sub:
                print(f"Manual reminder: Sending readiness to {name} ({email})...")
                send_deadline_reminder_email(email, name, "Semester Readiness", r_date_str)
                sent_count += 1
                
    # Closure manual send
    c_deadline = settings.get('closure_deadline')
    c_date_str = "N/A"
    if c_deadline:
        try:
            c_dt = datetime.strptime(c_deadline, "%Y-%m-%dT%H:%M")
            c_date_str = c_dt.strftime("%d %B %Y")
        except Exception as ex:
            print(f"Error parsing manual closure deadline: {ex}")
        
        for user in users_col.find():
            email = user.get('email')
            name = user.get('name')
            has_sub = submissions_col.find_one({
                'identity.submitterEmail': email,
                'form_type': 'closure',
                '_draft': False
            })
            if not has_sub:
                print(f"Manual reminder: Sending closure to {name} ({email})...")
                send_deadline_reminder_email(email, name, "Semester Closure", c_date_str)
                sent_count += 1
                
    return jsonify({'ok': True, 'sent_count': sent_count})

@erp_bp.route('/admin/send-targeted-reminders', methods=['POST'])
def send_targeted_reminders():
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})
    
    data = request.json or {}
    target_type = data.get('target_type')
    department = data.get('department')
    
    if not target_type:
        return jsonify({'ok': False, 'error': 'Missing target_type'})
        
    settings = get_global_settings()
    sent_count = 0
    
    r_deadline = settings.get('readiness_deadline')
    r_date_str = "N/A"
    if r_deadline:
        try:
            r_dt = datetime.strptime(r_deadline, "%Y-%m-%dT%H:%M")
            r_date_str = r_dt.strftime("%d %B %Y")
        except Exception as ex:
            print(f"Error parsing targeted readiness deadline: {ex}")
            
    c_deadline = settings.get('closure_deadline')
    c_date_str = "N/A"
    if c_deadline:
        try:
            c_dt = datetime.strptime(c_deadline, "%Y-%m-%dT%H:%M")
            c_date_str = c_dt.strftime("%d %B %Y")
        except Exception as ex:
            print(f"Error parsing targeted closure deadline: {ex}")
            
    # Helper to send readiness reminder if pending
    def check_and_send_readiness(user):
        nonlocal sent_count
        email = user.get('email')
        name = user.get('name')
        has_sub = submissions_col.find_one({
            'identity.submitterEmail': email,
            'form_type': 'readiness',
            '_draft': False
        })
        if not has_sub:
            print(f"Targeted reminder: Sending readiness to {name} ({email})...")
            if send_deadline_reminder_email(email, name, "Semester Readiness", r_date_str):
                sent_count += 1
                
    # Helper to send closure reminder if pending
    def check_and_send_closure(user):
        nonlocal sent_count
        email = user.get('email')
        name = user.get('name')
        has_sub = submissions_col.find_one({
            'identity.submitterEmail': email,
            'form_type': 'closure',
            '_draft': False
        })
        if not has_sub:
            print(f"Targeted reminder: Sending closure to {name} ({email})...")
            if send_deadline_reminder_email(email, name, "Semester Closure", c_date_str):
                sent_count += 1

    if target_type == 'all_pending':
        for user in users_col.find():
            check_and_send_readiness(user)
            check_and_send_closure(user)
            
    elif target_type == 'pending_readiness':
        for user in users_col.find():
            check_and_send_readiness(user)
            
    elif target_type == 'pending_closure':
        for user in users_col.find():
            check_and_send_closure(user)
            
    elif target_type == 'specific_dept':
        if not department:
            return jsonify({'ok': False, 'error': 'Missing department for specific_dept target group'})
            
        found_any = False
        for user in users_col.find():
            user_dept = user.get('department')
            if not user_dept:
                latest_sub = submissions_col.find_one({'identity.submitterEmail': user.get('email')}, sort=[('timestamp', -1)])
                if latest_sub and 'identity' in latest_sub:
                    user_dept = latest_sub['identity'].get('dept')
                    
            if user_dept == department:
                found_any = True
                check_and_send_readiness(user)
                check_and_send_closure(user)
                
        if not found_any:
            return jsonify({'ok': False, 'error': f'No registered HOD account found for department "{department}"'})
            
    else:
        return jsonify({'ok': False, 'error': f'Unknown target type "{target_type}"'})
        
    return jsonify({'ok': True, 'sent_count': sent_count})

@erp_bp.route('/admin/approve-edit/<sub_id>', methods=['POST'])
def approve_edit(sub_id):
    if not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Unauthorized'})

    submissions_col.update_one(
        {'_id': ObjectId(sub_id)},
        {'$set': {'edit_request.status': 'approved', 'edit_request.pending': False}}
    )
    return jsonify({'ok': True})

@erp_bp.route('/admin/export-all')
def export_all():
    if not session.get('admin'):
        return redirect(url_for('erp.admin_login'))
    docs = list(submissions_col.find().sort('timestamp', -1))
    for d in docs:
        fac_docs = list(faculty_submissions_col.find({'parent_submission_id': str(d['_id'])}))
        d['_faculty_submissions'] = fac_docs
    wb = build_workbook(docs)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(buf, as_attachment=True,
                     download_name=f"SemReadiness_ALL_{datetime.utcnow().strftime('%Y%m%d')}.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")

@erp_bp.route('/admin/delete/<sid>', methods=['POST'])
def delete_submission(sid):
    if not session.get('admin'):
        return jsonify({'ok': False})
    submissions_col.delete_one({'_id': ObjectId(sid)})
    # Also delete linked faculty submissions
    faculty_submissions_col.delete_many({'parent_submission_id': sid})
    return jsonify({'ok': True})

# Allow the HOD (owner) to also delete their own submission
@erp_bp.route('/api/delete-submission/<sid>', methods=['POST'])
def delete_my_submission(sid):
    if 'user_email' not in session and not session.get('admin'):
        return jsonify({'ok': False, 'error': 'Not logged in'})
    try:
        doc = submissions_col.find_one({'_id': ObjectId(sid)})
        if not doc:
            return jsonify({'ok': False, 'error': 'Not found'})
        # Owner-only unless admin
        if not session.get('admin') and doc.get('identity', {}).get('submitterEmail') != session.get('user_email'):
            return jsonify({'ok': False, 'error': 'Not authorised'})
        submissions_col.delete_one({'_id': ObjectId(sid)})
        faculty_submissions_col.delete_many({'parent_submission_id': sid})
        return jsonify({'ok': True})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)})


# ═════════════════════════════════════════════════════════════
# WORKBOOK BUILDERS
# ═════════════════════════════════════════════════════════════

def _styles():
    hdr_fill = PatternFill("solid", fgColor="0A2558")
    hdr_font = Font(color="FFFFFF", bold=True, size=10, name="Calibri")
    gold_fill = PatternFill("solid", fgColor="F4A819")
    gold_font = Font(color="1A1A1A", bold=True, size=10, name="Calibri")
    thin = Border(left=Side(style='thin'), right=Side(style='thin'),
                  top=Side(style='thin'), bottom=Side(style='thin'))
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return hdr_fill, hdr_font, gold_fill, gold_font, thin, center


def build_hod_checklist_only(doc, form_type='readiness'):
    """Build a single-sheet Excel that the HOD can download and share with faculty.
    Contains: identity header + HOD's filled checklist + a blank Faculty column for
    each faculty (or simply the checklist for reference).
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hdr_fill, hdr_font, gold_fill, gold_font, thin, center = _styles()

    label = 'Closure' if form_type == 'closure' else 'Readiness'
    ws = wb.create_sheet(f"HOD {label} Checklist")
    idt = doc.get('identity', {})

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=5)
    t = ws.cell(row=1, column=1, value=f"JAIN University — HOD Semester {label} Submission")
    t.font = Font(color="FFFFFF", bold=True, size=14, name="Calibri")
    t.fill = hdr_fill
    t.alignment = center
    ws.row_dimensions[1].height = 32

    # Identity block
    identity_rows = [
        ("Department",       idt.get('dept', '')),
        ("Campus",           idt.get('campus', '')),
        ("HOD Name",         idt.get('hodName', '') or idt.get('hod', '')),
        ("HOD Email",        idt.get('hodEmail', '') or idt.get('email', '')),
        ("Semester",         idt.get('semester', '') or idt.get('sem', '')),
        ("Academic Year",    idt.get('acYear', '')),
        ("Submission Date",  idt.get('subDate', '') or idt.get('date', '')),
        ("Faculty Deadline", idt.get('hodDeadline', '')),
        ("Mid Sem Review Date", doc.get('hodChecklist', {}).get('28', {}).get('status', '')),
    ]
    r = 3
    for label_, val in identity_rows:
        c1 = ws.cell(row=r, column=1, value=label_)
        c1.fill = gold_fill
        c1.font = gold_font
        c1.border = thin
        c1.alignment = Alignment(horizontal='left', vertical='center')
        c2 = ws.cell(row=r, column=2, value=val)
        c2.font = Font(name="Calibri", size=10)
        c2.border = thin
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
        r += 1

    r += 1

    # Checklist / Report headers
    if form_type == 'closure':
        headers = ['#', 'Vertical (HOD Report Section)', 'HOD Narrative', '', '']
    else:
        headers = ['#', 'Section', 'Checklist Item', 'Status', 'Remark / Notes']

    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=r, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin
    ws.row_dimensions[r].height = 28
    r += 1

    if form_type == 'closure':
        hod_rpt = doc.get('hodReport', {})
        for item in HOD_CLOSURE_SECTIONS:
            val = hod_rpt.get(item['id'], {})
            if isinstance(val, dict):
                row_vals = [item['id'], item['title'], val.get('text', ''), val.get('link', ''), val.get('file_url', '')]
            else:
                row_vals = [item['id'], item['title'], str(val), '', '']
            for c, v in enumerate(row_vals, 1):
                cell = ws.cell(row=r, column=c, value=v)
                cell.border = thin
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            r += 1
    else:
        chk = doc.get('hodChecklist', {})
        for sec in HOD_SECTIONS:
            sec_title = sec['title']
            for item in sec['items']:
                val = chk.get(item['id'], {})
                row_vals = [item['id'], sec_title, item['text'], val.get('status', ''), val.get('remark', '')]
                for c, v in enumerate(row_vals, 1):
                    cell = ws.cell(row=r, column=c, value=v)
                    cell.border = thin
                    cell.font = Font(name="Calibri", size=10)
                    cell.alignment = Alignment(wrap_text=True, vertical='top')
                    if c == 4:
                        if v == 'No':
                            cell.fill = PatternFill("solid", fgColor="FFE8E8")
                        elif v == 'Yes':
                            cell.fill = PatternFill("solid", fgColor="E8F8EE")
                r += 1

    for i, w in enumerate([8, 30, 55, 14, 38], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # ── Second sheet: a blank Faculty Checklist template (for faculty reference) ──
    _, fac_sections = _sections_for(form_type)
    ws2 = wb.create_sheet("Faculty Checklist (Template)")
    ws2.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    t2 = ws2.cell(row=1, column=1, value=f"Faculty Semester {label} Checklist — Auto-filled from HOD")
    t2.font = Font(color="FFFFFF", bold=True, size=12, name="Calibri")
    t2.fill = hdr_fill
    t2.alignment = center
    ws2.row_dimensions[1].height = 28

    fac_id_rows = [
        ("Department",     idt.get('dept', '')),
        ("Campus",         idt.get('campus', '')),
        ("HOD Name",       idt.get('hodName', '') or idt.get('hod', '')),
        ("Semester",       idt.get('semester', '') or idt.get('sem', '')),
        ("Academic Year",  idt.get('acYear', '')),
        ("Faculty Name",   ""),
        ("Faculty Email",  ""),
        ("Course Name",    ""),
        ("Course Code",    ""),
    ]
    r = 3
    for label_, val in fac_id_rows:
        c1 = ws2.cell(row=r, column=1, value=label_)
        c1.fill = gold_fill
        c1.font = gold_font
        c1.border = thin
        c2 = ws2.cell(row=r, column=2, value=val)
        c2.font = Font(name="Calibri", size=10)
        c2.border = thin
        ws2.merge_cells(start_row=r, start_column=2, end_row=r, end_column=4)
        r += 1
    r += 1

    headers2 = ['#', 'Faculty Checklist Item', 'Status (Yes/No/N/A)', 'Remarks / Evidence Link']
    for c, h in enumerate(headers2, 1):
        cell = ws2.cell(row=r, column=c, value=h)
        cell.font = hdr_font
        cell.fill = hdr_fill
        cell.alignment = center
        cell.border = thin
    ws2.row_dimensions[r].height = 28
    r += 1

    for sec in fac_sections:
        # section header row
        cell = ws2.cell(row=r, column=1, value=sec['title'])
        cell.fill = gold_fill
        cell.font = gold_font
        cell.border = thin
        cell.alignment = Alignment(horizontal='left')
        ws2.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
        r += 1
        for item in sec['items']:
            row_vals = [item['id'], item['text'], '', '']
            for c, v in enumerate(row_vals, 1):
                cell = ws2.cell(row=r, column=c, value=v)
                cell.border = thin
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            r += 1
    for i, w in enumerate([8, 60, 18, 36], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    return wb


def build_workbook(submissions):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    hdr_fill, hdr_font, gold_fill, gold_font, thin, center = _styles()

    def style_header(ws, headers, row=1):
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=c, value=h)
            cell.font = hdr_font
            cell.fill = hdr_fill
            cell.alignment = center
            cell.border = thin
        ws.row_dimensions[row].height = 36

    def add_title(ws, text, ncols):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        t = ws.cell(row=1, column=1, value=text)
        t.font = Font(color="FFFFFF", bold=True, size=12, name="Calibri")
        t.fill = hdr_fill
        t.alignment = center
        ws.row_dimensions[1].height = 28

    # Sheet 1 — Dashboard
    ws1 = wb.create_sheet("Dashboard")
    add_title(ws1, "JAIN University — Semester Submissions Dashboard", 20)
    heads = ['Timestamp', 'Form Type', 'Campus', 'Department', 'HOD Name', 'HOD Email', 'Semester', 'Acad Year', 'Sub Date', 'HOD Deadline', 'Mid Sem Date',
             'Faculty Submissions', 'HOD Checklist Yes/Total']
    style_header(ws1, heads, 2)
    for r, d in enumerate(submissions, 3):
        idt = d.get('identity', {})
        chk = d.get('hodChecklist', {})
        yes_count = sum(1 for v in chk.values() if isinstance(v, dict) and v.get('status') == 'Yes')
        fac_subs = len(d.get('_faculty_submissions', []))
        vals = [d.get('timestamp', ''), d.get('form_type', 'readiness'), idt.get('campus', ''),
                idt.get('dept', ''), idt.get('hodName', '') or idt.get('hod', ''),
                idt.get('hodEmail', '') or idt.get('email', ''),
                idt.get('semester', '') or idt.get('sem', ''), idt.get('acYear', ''),
                idt.get('subDate', '') or idt.get('date', ''), idt.get('hodDeadline', ''), chk.get('28', {}).get('status', ''),
                fac_subs, f"{yes_count}/{len(chk) if chk else 0}"]
        for c, v in enumerate(vals, 1):
            cell = ws1.cell(row=r, column=c, value=v)
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
            cell.fill = PatternFill("solid", fgColor="F5F8FF" if r % 2 == 0 else "FFFFFF")
    ws1.column_dimensions['A'].width = 20
    for i in range(2, 21):
        ws1.column_dimensions[get_column_letter(i)].width = 16

    # Sheet 2 — Programs
    ws2 = wb.create_sheet("Program Breakdown")
    add_title(ws2, "Program-wise Breakdown", 7)
    style_header(ws2, ['Department', 'Campus', 'Program Name', 'Courses', 'Students', 'Faculty', 'Coordinator'], 2)
    r = 3
    for d in submissions:
        for p in d.get('programs', []):
            vals = [d.get('identity', {}).get('dept', ''), d.get('identity', {}).get('campus', ''),
                    p.get('name', ''), p.get('courses', ''), p.get('students', ''),
                    p.get('faculty', ''), p.get('coord', '')]
            for c, v in enumerate(vals, 1):
                cell = ws2.cell(row=r, column=c, value=v)
                cell.border = thin
                cell.font = Font(name="Calibri", size=10)
            r += 1
    for i, w in enumerate([28, 14, 32, 10, 10, 10, 22], 1):
        ws2.column_dimensions[get_column_letter(i)].width = w

    # Sheet 3 — Faculty Submissions
    ws3 = wb.create_sheet("Faculty Submissions")
    add_title(ws3, "Faculty Submissions (via Shared Link)", 13)
    style_header(ws3, ['Form Type', 'Department', 'Campus', 'Semester', 'Faculty Name', 'Faculty Email',
                       'Course', 'Code', 'Program', 'Year/Sem', 'Students', 'HOD Remarks', 'Review Status'], 2)
    r = 3
    for d in submissions:
        for f in d.get('_faculty_submissions', []):
            vals = [f.get('form_type', d.get('form_type', 'readiness')),
                    f.get('dept', ''), f.get('campus', ''), f.get('semester', ''),
                    f.get('faculty_name', ''), f.get('faculty_email', ''),
                    f.get('course_name', ''), f.get('course_code', ''),
                    f.get('program', ''), f.get('year_sem', ''), f.get('no_of_students', ''),
                    f.get('hod_remarks', ''), f.get('hod_review_status', '')]
            for c, v in enumerate(vals, 1):
                cell = ws3.cell(row=r, column=c, value=v)
                cell.border = thin
                cell.font = Font(name="Calibri", size=10)
            r += 1
    for i, w in enumerate([12, 26, 12, 16, 22, 26, 24, 12, 18, 12, 10, 28, 14], 1):
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Sheet 4 — Faculty Checklist Detail
    ws4 = wb.create_sheet("Faculty Checklist Detail")
    add_title(ws4, "Faculty Checklist — Item-wise", 9)
    style_header(ws4, ['Form Type', 'Department', 'Faculty Name', 'Course', 'Code', 'Item #', 'Checklist Item', 'Status', 'Remark'], 2)
    r = 3
    for d in submissions:
        for f in d.get('_faculty_submissions', []):
            chk = f.get('checklist', {})
            ftype = f.get('form_type', d.get('form_type', 'readiness'))
            _, fac_sections = _sections_for(ftype)
            for sec in fac_sections:
                for item in sec['items']:
                    val = chk.get(item['id'], {})
                    if not isinstance(val, dict):
                        val = {}
                    vals = [ftype, f.get('dept', ''), f.get('faculty_name', ''), f.get('course_name', ''),
                            f.get('course_code', ''), item['id'], item['text'],
                            val.get('status', ''), val.get('remark', '')]
                    for c, v in enumerate(vals, 1):
                        cell = ws4.cell(row=r, column=c, value=v)
                        cell.border = thin
                        cell.font = Font(name="Calibri", size=10)
                        if val.get('status') == 'No':
                            cell.fill = PatternFill("solid", fgColor="FFE8E8")
                        elif val.get('status') == 'Yes':
                            cell.fill = PatternFill("solid", fgColor="E8F8EE")
                    r += 1
    for i, w in enumerate([12, 26, 22, 24, 12, 8, 48, 10, 30], 1):
        ws4.column_dimensions[get_column_letter(i)].width = w

    # Sheet 5 — HOD Checklist (Readiness)
    ws5 = wb.create_sheet("HOD Checklist")
    add_title(ws5, "HOD Checklist — All Departments", 8)
    style_header(ws5, ['Department', 'Campus', 'Semester', 'Item #', 'Checklist Item', 'Section', 'Status', 'Remark'], 2)
    r = 3
    for d in submissions:
        if d.get('form_type', 'readiness') != 'readiness':
            continue
        chk = d.get('hodChecklist', {})
        for sec in HOD_SECTIONS:
            for item in sec['items']:
                val = chk.get(item['id'], {})
                if not isinstance(val, dict):
                    val = {}
                vals = [d.get('identity', {}).get('dept', ''), d.get('identity', {}).get('campus', ''),
                        d.get('identity', {}).get('semester', ''), item['id'], item['text'],
                        sec['title'].replace('Section ', '').split(': ', 1)[-1],
                        val.get('status', ''), val.get('remark', '')]
                for c, v in enumerate(vals, 1):
                    cell = ws5.cell(row=r, column=c, value=v)
                    cell.border = thin
                    cell.font = Font(name="Calibri", size=10)
                    if val.get('status') == 'No':
                        cell.fill = PatternFill("solid", fgColor="FFE8E8")
                    elif val.get('status') == 'Yes':
                        cell.fill = PatternFill("solid", fgColor="E8F8EE")
                r += 1
    for i, w in enumerate([26, 12, 14, 8, 52, 24, 10, 30], 1):
        ws5.column_dimensions[get_column_letter(i)].width = w

    # Sheet 6 — HOD Closure Report
    ws_c = wb.create_sheet("HOD Closure Reports")
    add_title(ws_c, "HOD Semester Closure — Narrative Reports", 5)
    style_header(ws_c, ['Department', 'Campus', 'Semester', 'Vertical', 'HOD Narrative'], 2)
    r = 3
    for d in submissions:
        if d.get('form_type') != 'closure':
            continue
        rpt = d.get('hodReport', {})
        idt = d.get('identity', {})
        for item in HOD_CLOSURE_SECTIONS:
            val = rpt.get(item['id'], {})
            if isinstance(val, dict):
                text = val.get('text', '')
                link = val.get('link', '')
                file_url = val.get('file_url', '')
                if link or file_url:
                    parts = [text]
                    if link:
                        parts.append(f"[Link: {link}]")
                    if file_url:
                        parts.append(f"[File: {file_url}]")
                    text = " \n".join(parts)
            else:
                text = str(val)
            vals = [idt.get('dept', ''), idt.get('campus', ''),
                    idt.get('semester', '') or idt.get('sem', ''),
                    f"{item['id']}. {item['title']}", text]
            for c, v in enumerate(vals, 1):
                cell = ws_c.cell(row=r, column=c, value=v)
                cell.border = thin
                cell.font = Font(name="Calibri", size=10)
                cell.alignment = Alignment(wrap_text=True, vertical='top')
            r += 1
    for i, w in enumerate([26, 12, 14, 38, 80], 1):
        ws_c.column_dimensions[get_column_letter(i)].width = w

    # Sheet 7 — Gap Analysis (Readiness only)
    ws6 = wb.create_sheet("Gap Analysis")
    add_title(ws6, "Gap Analysis — Readiness Checklist Completion by Department", 9)
    style_header(ws6, ['Department', 'Campus', 'Semester', 'Yes', 'No', 'N/A', 'Blank', 'Completion %', 'Gap Items (No)'], 2)
    all_items = [item for sec in HOD_SECTIONS for item in sec['items']]
    r = 3
    for d in submissions:
        if d.get('form_type', 'readiness') != 'readiness':
            continue
        chk = d.get('hodChecklist', {})
        vals_list = [v for v in chk.values() if isinstance(v, dict)]
        yes = sum(1 for x in vals_list if x.get('status') == 'Yes')
        no = sum(1 for x in vals_list if x.get('status') == 'No')
        na = sum(1 for x in vals_list if x.get('status') == 'N/A')
        blank = sum(1 for x in vals_list if not x.get('status'))
        denom = (yes + no + blank)
        pct = f"{round(yes / denom * 100)}%" if denom else "—"
        gaps = "; ".join(f"{i['id']}. {i['text']}" for i in all_items
                         if isinstance(chk.get(i['id']), dict) and chk[i['id']].get('status') == 'No')
        row_vals = [d.get('identity', {}).get('dept', ''), d.get('identity', {}).get('campus', ''),
                    d.get('identity', {}).get('semester', ''), yes, no, na, blank, pct, gaps]
        for c, v in enumerate(row_vals, 1):
            cell = ws6.cell(row=r, column=c, value=v)
            cell.border = thin
            cell.font = Font(name="Calibri", size=10)
            if c == 5 and isinstance(v, int) and v > 0:
                cell.fill = PatternFill("solid", fgColor="FFE8E8")
                cell.font = Font(name="Calibri", size=10, bold=True, color="AA0000")
        r += 1
    for i, w in enumerate([26, 12, 14, 8, 8, 8, 8, 14, 60], 1):
        ws6.column_dimensions[get_column_letter(i)].width = w

    return wb



@erp_bp.route('/api/track-install', methods=['POST'])
def track_install():
    db['pwa_analytics'].update_one(
        {'_id': 'pwa_stats'},
        {'$inc': {'installs': 1}},
        upsert=True
    )
    return jsonify({'ok': True})

@erp_bp.route('/api/track-launch', methods=['POST'])
def track_launch():
    db['pwa_analytics'].update_one(
        {'_id': 'pwa_stats'},
        {'$inc': {'launches': 1}},
        upsert=True
    )
    return jsonify({'ok': True})

